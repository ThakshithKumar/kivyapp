import kivy
from kivymd.app import MDApp
from kivy.properties import StringProperty
from kivymd.uix.label import MDLabel
from kivy.uix.gridlayout import GridLayout
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.lang import Builder
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.widget import Widget
from kivy_garden.zbarcam import ZBarCam
import qrcode
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter


class MainWindow(Screen):
    pass


class FirstWindow(Screen):
    pass


class SecondWindow(Screen):
    pass


class WindowManager(ScreenManager):
    pass


class ExcelData:
    def accessing_data(self):
        wb = load_workbook('Book1.xlsx')
        ws = wb.active
        for row in range(1, 5):
            for col in range(1, 4):
                char = get_column_letter(col)
                print(ws[char + str(row)].value)

class GeneratorWindow(Screen):
    pass


text1 = "Partially Vaccinated!!"
text2 = "Fully Vaccinated!!"


class Function(Screen):
    def generate_qrcode(self, root):
        code = qrcode.QRCode(version=None)
        code.add_data(text1)
        code.make(fit=True)
        img = code.make_image()
        img.save("image.png")

    def generate_qrcode2(self, root):
        code = qrcode.QRCode(version=None)
        code.add_data(text2)
        code.make(fit=True)
        img = code.make_image()
        img.save("image2.png")



class Dose1Window(Screen, Widget):
    pass


class Dose2Window(Screen):
    pass


class ScannerWindow(Screen):
    pass


class QRapp(MDApp):
    def build(self):
        self.root = Builder.load_file("my.kv")

    #def welcome(self):
        #self.root.ids.welcom.text = f'Welcome {self.root.ids.usernm.text}'


if __name__ == "__main__":
    QRapp().run()
